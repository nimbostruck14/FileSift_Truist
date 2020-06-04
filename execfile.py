#This is a program that takes in a pdf document and extracts data into a formatted excel table
#In order for this to work, the pdf and the excel sheet must be put into the working directory folder
#-------------------------------------------------
# External Modules:
import PyPDF2
import openpyxl as xl
#-------------------------------------------------

#================================================
#                   Rev Notes:
# This current rendition still needs to have Coverage,
# Spread Changes, and Pull Through Marketing fixed
# Additionally, writing to an excel file and correctly
# extracting data from a pdf file

#================================================

# pdffile = open(input("Enter the file name and extension for data extraction: "), "rb")
# pdfread = PyPDF2.PdfFileReader(pdffile)
# pagenum = pdfread.numPages      #number of pages
# pageobj = pdfread.getPage(pagenum-1)
#
# txtdata = pageobj.extractText()     #stores extracted text data
#
# txtfile = open(r"C:\Users\Nick's Console\PycharmProjects\FileSift_Truist\\text.txt", "a")
# txtfile.writelines(txtdata)
# pdffile.close()
# txtfile.close()
# siftfile = open('text.txt', "r")

siftfile = open(input("Enter the file name and extension for data extraction: "), "r")


#-------read file contents----------------
readfile = siftfile.readlines()
print(readfile)

#-----------Word Functions----------------
def tsmpl():
    index = 0
    dex = 0
    out = ['0']
    write_flag = False
    param = 'Total Secondary Marketing P/L'
    pl = len(param)
    dl = len(readfile)
    #acquires Total Secondary Marketing P/L
    for index in range(dl):
        transf_obj = readfile[index]
        if transf_obj[0:pl] == param:
            transf_obj_list = list(transf_obj)
            for dex in range(len(transf_obj_list)):
                if transf_obj_list[dex] == '0' or transf_obj_list[dex] == '1' or transf_obj_list[dex] == '2' or \
                        transf_obj_list[dex] == '3' or transf_obj_list[dex] == '4' or transf_obj_list[dex] == '5' or \
                        transf_obj_list[dex] == '6' or transf_obj_list[dex] == '7' or transf_obj_list[dex] == '8' or \
                        transf_obj_list[dex] == '9' or transf_obj_list[dex] == '(' or transf_obj_list[dex] == ')' or \
                        transf_obj_list[dex] == ',':
                    write_flag = True
                else:
                    if write_flag:
                        del out[0]
                        return out
                    else:
                        write_flag = False
                if write_flag:
                    out.append(transf_obj_list[dex])
                dex += 1
            return out
        index += 1

def coverage():
    index = 0
    dex = 0
    dext = 0
    out = ['0']
    write_flag = False
    param = 'Coverage'
    pl = len(param)
    dl = len(readfile)
    #acquires Pull-Through Model
    for index in range(dl):                         #Goes through entire file
        transf_obj = readfile[index]                #
        if transf_obj[0:pl] == param:
            transf_obj_list = list(readfile[index])
            for dext in range(len(transf_obj_list)):
                if transf_obj_list[dext] == 'C' and transf_obj_list[dext+1] == 'o':
                    dex = dext
                    while dex in range(len(transf_obj_list)):
                        if transf_obj_list[dex] == '0' or transf_obj_list[dex] == '1' or transf_obj_list[dex] == '2' or \
                                transf_obj_list[dex] == '3' or transf_obj_list[dex] == '4' or transf_obj_list[dex] == '5' or \
                                transf_obj_list[dex] == '6' or transf_obj_list[dex] == '7' or transf_obj_list[dex] == '8' or \
                                transf_obj_list[dex] == '9' or transf_obj_list[dex] == '(' or transf_obj_list[dex] == ')' or \
                                transf_obj_list[dex] == ',':
                            write_flag = True
                        else:
                            if write_flag:
                                del out[0]
                                return out
                            else:
                                write_flag = False
                        if write_flag:
                            out.append(transf_obj_list[dex])
                        dex += 1
                    return out
        index += 1

def exposure():
    index = 0
    dex = 0
    out = ['0']
    write_flag = False
    param = 'Exposure'
    pl = len(param)
    dl = len(readfile)
    #acquires Total Secondary Marketing P/L
    for index in range(dl):
        transf_obj = readfile[index]
        if transf_obj[0:pl] == param:
            transf_obj_list = list(transf_obj)
            for dex in range(len(transf_obj_list)):
                if transf_obj_list[dex] == '0' or transf_obj_list[dex] == '1' or transf_obj_list[dex] == '2' or \
                        transf_obj_list[dex] == '3' or transf_obj_list[dex] == '4' or transf_obj_list[dex] == '5' or \
                        transf_obj_list[dex] == '6' or transf_obj_list[dex] == '7' or transf_obj_list[dex] == '8' or \
                        transf_obj_list[dex] == '9' or transf_obj_list[dex] == '(' or transf_obj_list[dex] == ')' or \
                        transf_obj_list[dex] == ',':
                    write_flag = True
                else:
                    if write_flag:
                        del out[0]
                        return out
                    else:
                        write_flag = False
                if write_flag:
                    out.append(transf_obj_list[dex])
                dex += 1
            return out
        index += 1

def pfm():
    index = 0
    dex = 0
    dext = 0
    out = ['0']
    write_flag = False
    param = 'Pull-Through Model'
    pl = len(param)
    dl = len(readfile)
    #acquires Pull-Through Model
    for index in range(dl):                         #Goes through entire file
        transf_obj = readfile[index]                #
        if param in transf_obj:
            transf_obj_list = list(readfile[index])
            for dext in range(len(transf_obj_list)):
                if transf_obj_list[dext] == 'P':
                    dex = dext
                    while dex in range(len(transf_obj_list)):
                        if transf_obj_list[dex] == '0' or transf_obj_list[dex] == '1' or transf_obj_list[dex] == '2' or \
                                transf_obj_list[dex] == '3' or transf_obj_list[dex] == '4' or transf_obj_list[dex] == '5' or \
                                transf_obj_list[dex] == '6' or transf_obj_list[dex] == '7' or transf_obj_list[dex] == '8' or \
                                transf_obj_list[dex] == '9' or transf_obj_list[dex] == '(' or transf_obj_list[dex] == ')' or \
                                transf_obj_list[dex] == ',':
                            write_flag = True
                        else:
                            if write_flag:
                                del out[0]
                                return out
                            else:
                                write_flag = False
                        if write_flag:
                            out.append(transf_obj_list[dex])
                        dex += 1
                    return out
        index += 1

def spread():
    index = 0
    dex = 0
    dext = 0
    out = ['0']
    write_flag = False
    param = 'Spread Changes'
    pl = len(param)
    dl = len(readfile)
    #acquires Pull-Through Model
    for index in range(dl):                         #Goes through entire file
        transf_obj = readfile[index]                #@ 70
        if param in transf_obj:
            transf_obj_list = list(readfile[index])
            for dext in range(len(transf_obj_list)):
                if transf_obj_list[dext] == 'S' and transf_obj_list[dext+1] == 'p':
                    dex = dext
                    while dex in range(len(transf_obj_list)):
                        if transf_obj_list[dex] == '0' or transf_obj_list[dex] == '1' or transf_obj_list[dex] == '2' or \
                                transf_obj_list[dex] == '3' or transf_obj_list[dex] == '4' or transf_obj_list[dex] == '5' or \
                                transf_obj_list[dex] == '6' or transf_obj_list[dex] == '7' or transf_obj_list[dex] == '8' or \
                                transf_obj_list[dex] == '9' or transf_obj_list[dex] == '(' or transf_obj_list[dex] == ')' or \
                                transf_obj_list[dex] == ',':
                            write_flag = True
                        else:
                            if write_flag:
                                del out[0]
                                return out
                            else:
                                write_flag = False
                        if write_flag:
                            out.append(transf_obj_list[dex])
                        dex += 1
                    return out
        index += 1


#-------------------------------------------------

#opens text file

#-------------------------------------------------
# Methods and code block to read the pdf
#-------------------------------------------------

#-------Printing Object Blocks--------------------
tsmplout = tsmpl()
tsmplobj = ''.join(tsmplout)
print("tsmpl:")
if tsmplobj == "-":
    tsmplobj = "0"
print(tsmplobj)

covout = coverage()
covobj = ''.join(covout)
print("Coverage:")
if covobj == "-":
    covobj = "0"
print(covobj)

expout = exposure()
expobj = ''.join(expout)
print("Exposure:")
if expobj == "-":
    expobj = "0"
print(expobj)

spdout = spread()
spdobj = ''.join(spdout)
print("Spread Changes:")
if spdobj == "-":
    spdobj = "0"
print(spdobj)

pfmout = pfm()
pfmobj = ''.join(pfmout)
print("Pull-Through Model:")
if pfmobj == "-":
    pfmobj = "0"
print(pfmobj)

#print(pfmobj)



#-------------------------------------------------

#closes the read file
siftfile.close()

#Opens an excel document to write to
filename1 ="C:\\Users\\Nick's Console\\PycharmProjects\\FileSift_Truist\\Comb.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

#-------------------------------------------------
# Methods and code block to writing to spreadsheet
#-------------------------------------------------

ws2.cell(row = 2, column = 6).value = tsmplobj
ws2.cell(row = 2, column = 5).value = covobj
ws2.cell(row = 2, column = 2).value = expobj
ws2.cell(row = 2, column = 4).value = spdobj
ws2.cell(row = 2, column = 3).value = pfmobj

# saving the destination excel file
wb2.save(str(filename1))

#closes the write file
#writefile.close()